import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import validators

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Get the active sheet
sheet = workbook.active

# Set the headers for the sheet
sheet.cell(row=1, column=1).value = "Product Name"
sheet.cell(row=1, column=2).value = "Description"
sheet.cell(row=1, column=3).value = "Price"
sheet.cell(row=1, column=4).value = "Product link"

# To parse content for each page
def websitecontent(page_num):
    response = requests.get(f"https://getpalma.com/collections/all-bags?page={page_num}")

    # Creat soup object to parse content
    soup = BeautifulSoup(response.text, "html.parser")

    # Find all the product elements on the page
    products = soup.find_all("div", class_="product-block__inner")
    return products
#Add
# To download Images for each product then save the downloaded images in a folder names as the product name
def download_images(product_link, product_name):
    # Create a new folder for the product
    os.makedirs(product_name, exist_ok=True)

    # Send a GET request to the product page
    response = requests.get(product_link)

    # Parse the HTML content of the product page
    soup = BeautifulSoup(response.content, "html.parser")

    # Find all the image elements on the page
    images = soup.find_all("img")
    imageLinks=[]
    for image in images:
        image_link = image.attrs.get("data-src")
        if not image_link:
            continue
        else:
            #print(image_link)
            finalLink = "https:"+image_link
            valid = validators.url(finalLink)
            if valid == True:
                imageLinks.append(finalLink)
                print("Url is valid")
            else:
                print("Invalid url")

    counter = 1
    # Iterate through each image element
    for image in imageLinks:
        # Send a GET request to the image URL
        response = requests.get(image)
        # Get the image file name
        #image_name = os.path.basename("image_"+i+".JPG")
        #image_name = image_link.split("/")[-1]
        ext = ".jpg"
        # Save the image to the product folder
        open(f"{product_name}/{counter}{ext}", "wb").write(response.content)
        print(f"{counter}{ext} saved")
        counter+=1

# To parse Decription for each product by taking the product link as an input
# Define the productDescription function
def productDescription(link):
    response = requests.get(link)
    soup = BeautifulSoup(response.text, "html.parser")
    div_element = soup.find("div", id="smart-tabs-content-0")
    if div_element:
        description = div_element.text.strip()
    else:
        description = "Description not found"
    return description

# Number of pages in website
pages = 7

# Set the starting row for the data
row = 2

# Loop on all the pages in Bags section
for page in range(pages):
    products = websitecontent(page)
    # Iterate through each product element
    for product in products:
        # Find the product name
        product_name = product.find("span", class_="title").text

        # Find the product price
        price = product.find("span", class_="money").text

        # Find the product Link
        product_link = product.find("div", class_= "cc-quick-buy-btn-container")
        link=('https://getpalma.com/'+product_link.find("a").attrs['href'])

        print(link)

        # Find the product description
        description = productDescription(link)

        # Download prodcut images
        download = download_images(link, product_name)


        # Write the data to the sheet
        sheet.cell(row=row, column=1).value = product_name
        sheet.cell(row=row, column=2).value = description
        sheet.cell(row=row, column=3).value = price
        sheet.cell(row=row, column=4).value = link
        row += 1

    # Save the workbook
    workbook.save("products.xlsx")


